"""XLSX extraction adapter, wrapping ``openpyxl`` (extra ``viparse[office]``).

Each worksheet becomes a heading block (the sheet name) followed by a table block
(its cells), so S4 renders one GFM table per sheet. Every cell's font name is
recorded as a signal for the S3 detector — the same mechanism the DOCX engine uses.
The engine applies no Vietnamese logic itself.

``openpyxl`` is imported lazily inside :meth:`XlsxEngine.extract`, so importing this
module never requires the dependency; only extraction does. Legacy binary ``.xls``
is a separate engine (M3).
"""

from __future__ import annotations

import datetime
from typing import Any

from viparse.detect import CONTENT_TYPE_XLSX
from viparse.engines._shared import blocks_to_text
from viparse.errors import MissingDependency
from viparse.model import RawExtraction
from viparse.options import LoadOptions
from viparse.protocols import DEFAULT_PRIORITY, Source

_INSTALL_HINT = (
    "openpyxl is required for XLSX extraction; install it with: pip install 'viparse[office]'"
)


def _import_openpyxl() -> Any:
    """Import ``openpyxl`` lazily, raising a clear error if it is missing."""
    try:
        import openpyxl
    except ImportError as exc:
        raise MissingDependency(_INSTALL_HINT) from exc
    return openpyxl


class XlsxEngine:
    """Extracts each worksheet of a ``.xlsx`` file as a heading + table block."""

    priority = DEFAULT_PRIORITY
    #: Dependency + extra reported by ``viparse doctor`` (shared with the DOCX engine).
    dependency = "openpyxl"
    extra = "office"

    def supports(self, content_type: str) -> bool:
        return content_type == CONTENT_TYPE_XLSX

    def extract(self, source: Source, options: LoadOptions) -> RawExtraction:
        openpyxl = _import_openpyxl()
        formulas = openpyxl.load_workbook(str(source), data_only=False)
        # A second (cached-values) load is only useful for formula cells, so skip it —
        # and its full re-parse — for the common formula-free workbook.
        if _has_formula_cell(formulas):
            values = openpyxl.load_workbook(str(source), data_only=True)
        else:
            values = formulas
        blocks: list[dict[str, Any]] = []
        fonts: set[str] = set()
        warnings: list[str] = []
        for name in values.sheetnames:
            rows, sheet_fonts, uncomputed = _sheet_rows(values[name], formulas[name])
            fonts.update(sheet_fonts)
            if uncomputed:
                warnings.append(
                    f"sheet {name!r}: formula cell(s) had no cached value; kept the formula text"
                )
            if not rows:
                continue  # a blank sheet contributes no blocks
            blocks.append({"type": "heading", "level": 1, "text": name})
            blocks.append({"type": "table", "rows": rows})
        # Record the sheet on metadata only when unambiguous (a single-sheet workbook);
        # otherwise each sheet name lives in its heading block.
        sheet = values.sheetnames[0] if len(values.sheetnames) == 1 else None
        return RawExtraction(
            source=str(source),
            content_type=CONTENT_TYPE_XLSX,
            text=blocks_to_text(blocks),
            engine="xlsx",
            sheet=sheet,
            signals={"fonts": sorted(fonts), "blocks": blocks},
            warnings=warnings,
        )


def _sheet_rows(value_sheet: Any, formula_sheet: Any) -> tuple[list[list[str]], set[str], bool]:
    """Read a worksheet into rows of cell strings, collecting fonts and formula fallbacks.

    ``value_sheet`` carries cached computed values; ``formula_sheet`` carries the raw
    formula text. A cell with no cached value that is a formula keeps its formula text
    (and flags the sheet), per SPEC-2 T2.3.3. Merged cells surface as the value in the
    anchor cell and empty strings elsewhere — the natural spreadsheet grid.
    """
    rows: list[list[str]] = []
    fonts: set[str] = set()
    uncomputed = False
    # The two workbooks are the same file, so their row/column shapes match exactly.
    for value_row, formula_row in zip(
        value_sheet.iter_rows(), formula_sheet.iter_rows(), strict=True
    ):
        cells: list[str] = []
        for value_cell, formula_cell in zip(value_row, formula_row, strict=True):
            if formula_cell.font.name:
                fonts.add(formula_cell.font.name)
            value = value_cell.value
            if value is None and formula_cell.data_type == "f":
                value = formula_cell.value  # cached value absent → keep the formula text
                uncomputed = True
            cells.append(_cell_text(value))
        rows.append(cells)
    return _trim_trailing_empty_rows(rows), fonts, uncomputed


def _has_formula_cell(workbook: Any) -> bool:
    """Return ``True`` if any cell in the workbook is a formula."""
    return any(
        cell.data_type == "f"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def _cell_text(value: object) -> str:
    """Render a cell value as text the way a spreadsheet would show it.

    Plain ``str`` would leak Python artifacts — a midnight timestamp on date cells,
    ``True``/``False`` instead of Excel's ``TRUE``/``FALSE``, and IEEE-754 noise from
    cached float results — so those types are formatted explicitly.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        return format(value, ".12g")  # trims binary rounding noise; integral floats → "30"
    if isinstance(value, datetime.datetime):
        # openpyxl returns datetime even for date-only cells; drop a midnight time.
        return (
            value.date().isoformat()
            if value.time() == datetime.time()
            else value.isoformat(sep=" ")
        )
    if isinstance(value, datetime.date | datetime.time):
        return value.isoformat()
    return str(value)


def _trim_trailing_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    """Drop trailing all-empty rows (openpyxl over-reports the used range)."""
    while rows and all(cell == "" for cell in rows[-1]):
        rows.pop()
    return rows
