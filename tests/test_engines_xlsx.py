"""Tests for the XLSX extraction adapter."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

from viparse.detect import CONTENT_TYPE_XLSX
from viparse.engines.xlsx import XlsxEngine, _trim_trailing_empty_rows
from viparse.errors import MissingDependency
from viparse.options import LoadOptions
from viparse.registry import EngineRegistry

openpyxl = pytest.importorskip("openpyxl")  # skipped if the office extra is absent


def _make_xlsx(path: Path) -> Path:
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Tên"
    ws["B1"] = "Tuổi"
    ws["A2"] = "An"
    ws["B2"] = 30
    ws["A1"].font = Font(name=".VnTime")  # a legacy-font signal for S3
    wb.save(str(path))
    return path


def test_extract_returns_raw_extraction(tmp_path: Path) -> None:
    raw = XlsxEngine().extract(_make_xlsx(tmp_path / "a.xlsx"), LoadOptions())
    assert raw.engine == "xlsx"
    assert raw.content_type == CONTENT_TYPE_XLSX
    assert "Tên" in raw.text
    assert "An\t30" in raw.text  # a table row, tab-separated


def test_sheet_becomes_heading_plus_table(tmp_path: Path) -> None:
    raw = XlsxEngine().extract(_make_xlsx(tmp_path / "a.xlsx"), LoadOptions())
    blocks = raw.signals["blocks"]
    assert [b["type"] for b in blocks] == ["heading", "table"]
    assert blocks[0] == {"type": "heading", "level": 1, "text": "Data"}
    assert blocks[1]["rows"] == [["Tên", "Tuổi"], ["An", "30"]]


def test_single_sheet_name_is_recorded_on_metadata(tmp_path: Path) -> None:
    raw = XlsxEngine().extract(_make_xlsx(tmp_path / "a.xlsx"), LoadOptions())
    assert raw.sheet == "Data"


def test_captures_cell_font_signal(tmp_path: Path) -> None:
    raw = XlsxEngine().extract(_make_xlsx(tmp_path / "a.xlsx"), LoadOptions())
    assert ".VnTime" in raw.signals["fonts"]


def test_multiple_sheets_each_become_a_section(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "One"
    wb.active["A1"] = "x"
    second = wb.create_sheet("Two")
    second["A1"] = "y"
    path = tmp_path / "multi.xlsx"
    wb.save(str(path))
    raw = XlsxEngine().extract(path, LoadOptions())
    kinds = [(b["type"], b.get("text")) for b in raw.signals["blocks"]]
    assert kinds == [("heading", "One"), ("table", None), ("heading", "Two"), ("table", None)]
    assert raw.sheet is None  # ambiguous with more than one sheet


def test_merged_cells_surface_value_in_anchor_only(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "merged"
    ws.merge_cells("A1:B1")
    ws["A2"] = "left"
    ws["B2"] = "right"
    path = tmp_path / "merged.xlsx"
    wb.save(str(path))
    raw = XlsxEngine().extract(path, LoadOptions())
    table = next(b for b in raw.signals["blocks"] if b["type"] == "table")
    assert table["rows"] == [["merged", ""], ["left", "right"]]


def test_uncomputed_formula_keeps_text_and_warns(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=A1+A2"  # openpyxl writes no cached value → data_only sees None
    path = tmp_path / "formula.xlsx"
    wb.save(str(path))
    raw = XlsxEngine().extract(path, LoadOptions())
    table = next(b for b in raw.signals["blocks"] if b["type"] == "table")
    assert table["rows"] == [["10"], ["20"], ["=A1+A2"]]
    assert any("formula" in w for w in raw.warnings)


def test_blank_sheet_contributes_no_blocks(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()  # a fresh workbook has one empty sheet
    path = tmp_path / "empty.xlsx"
    wb.save(str(path))
    raw = XlsxEngine().extract(path, LoadOptions())
    assert raw.signals["blocks"] == []
    assert raw.text == ""


def test_cell_without_a_font_name_is_skipped(tmp_path: Path) -> None:
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["A1"].font = Font(name=None)  # a cell with no font name contributes no signal
    path = tmp_path / "nofont.xlsx"
    wb.save(str(path))
    raw = XlsxEngine().extract(path, LoadOptions())
    assert raw.signals["fonts"] == []
    assert raw.signals["blocks"][1]["rows"] == [["x"]]


def test_cell_text_formats_types_like_a_spreadsheet() -> None:
    import datetime

    from viparse.engines.xlsx import _cell_text

    assert _cell_text(None) == ""
    assert _cell_text(True) == "TRUE"
    assert _cell_text(False) == "FALSE"
    assert _cell_text(0.1 + 0.2) == "0.3"  # no IEEE-754 noise
    assert _cell_text(30.0) == "30"  # integral float
    assert _cell_text(datetime.datetime(2024, 1, 15)) == "2024-01-15"  # midnight → date only
    assert _cell_text(datetime.datetime(2024, 1, 15, 9, 30)) == "2024-01-15 09:30:00"
    assert _cell_text(datetime.date(2024, 1, 15)) == "2024-01-15"
    assert _cell_text(datetime.time(9, 30)) == "09:30:00"
    assert _cell_text("=A1+B1") == "=A1+B1"
    assert _cell_text(42) == "42"


def test_boolean_cell_renders_uppercase(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active["A1"] = True
    path = tmp_path / "bool.xlsx"
    wb.save(str(path))
    raw = XlsxEngine().extract(path, LoadOptions())
    assert raw.signals["blocks"][1]["rows"] == [["TRUE"]]


def test_trims_trailing_empty_rows() -> None:
    assert _trim_trailing_empty_rows([["a"], [""], ["", ""]]) == [["a"]]
    assert _trim_trailing_empty_rows([]) == []


def test_supports_only_xlsx() -> None:
    engine = XlsxEngine()
    assert engine.supports(CONTENT_TYPE_XLSX)
    assert not engine.supports("application/pdf")


def test_registry_selects_xlsx_engine() -> None:
    reg = EngineRegistry()
    reg.register(XlsxEngine())
    assert isinstance(reg.select(CONTENT_TYPE_XLSX), XlsxEngine)


def test_missing_dependency_raises_clear_error(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setitem(sys.modules, "openpyxl", None)
    with pytest.raises(MissingDependency, match=r"viparse\[office\]"):
        XlsxEngine().extract("missing.xlsx", LoadOptions())
